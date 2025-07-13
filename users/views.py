from django.contrib.auth import authenticate
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.utils import timezone
from datetime import timedelta
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, filters
from rest_framework.views import APIView
from rest_framework import status
from django.forms.models import model_to_dict
from django.db.models import Q
from rest_framework.pagination import PageNumberPagination
from decimal import Decimal
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from django.utils.encoding import smart_str
from rest_framework.parsers import MultiPartParser
from django.conf import settings
import os
from io import BytesIO
from openpyxl import load_workbook, Workbook
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import Product, Category

import random
import json
import threading
import os

from .models import CustomUser, ExpireToken, OTP, CustomUserRole, Product, Order, OrderItem, Payment, Category
from .utils import send_email, is_user_exist, threaded_send_email, parse_boolean, sanitize_price, sanitize_string, sanitize_tag
from .serializers import ProductSerializer
from .authentication import ExpireTokenAuthentication


# @api_view(['POST'])
# def login(request):
#     username = request.data.get('username')
#     password = request.data.get('password')
    
#     user = authenticate(request, username = username, password = password)
    
#     print(user, username, password)
  
#     if user is None:
#         return Response({"message": "Invalid Credentials"})
    
#     ExpireToken.objects.filter(user=user).delete()
#     token = ExpireToken.objects.create(user=user)
    
#     return Response({"Token": token.key})

@api_view(['POST'])
def login(request):
    username = request.data.get('username')
    password = request.data.get('password')

    print("Login attempt:", username, password)

    user = authenticate(username=username, password=password)

    if user is None:
        print("Authentication failed")
        return Response({"message": "Invalid Credentials"})

    print("Authenticated user:", user)

    ExpireToken.objects.filter(user=user).delete()
    token = ExpireToken.objects.create(user=user)

    return Response({"Token": token.key})

@api_view(["POST"])
def register(request):
    email = request.data.get("email")
    username = request.data.get("username")
    password = request.data.get("password")

    if is_user_exist(email):
        return Response({"Message": "Email is Already Registerd"})
    
    try:
        otp = OTP.objects.get(email = email)
    except OTP.DoesNotExist:
        return Response({"message": "Invalid Email"})
    
    if not otp.is_verified:
        return Response({"Message": "Email is not Verified"})
    
    try:
        CustomUser.objects.create(
            username = username,
            email = email,
            password = password,
        )
    except IntegrityError as e:
        transaction.rollback()
        
        if "username" in str(e):
            return Response({'Error': "username is Already Exist"})
        else: 
            return Response({'Error': f"Details: {e}"})
    
    otp.delete()

    return Response({"Message": "User Registered Succesfully"})

    
@api_view(['GET'])
def send_mail(request):
    email = request.GET.get('email')
    
    try:
        validate_email(email)
    except ValidationError:
        return {"message": "Email is not Valid"}
    
    if is_user_exist(email):
        return Response({"Message": "Email is Already Registerd"})
    
    try:
        otp_sent = OTP.objects.get(email = email)
    except OTP.DoesNotExist:
        otp_sent = None
    
    if otp_sent and (otp_sent.expire - timedelta(minutes = 4)) > timezone.now():
        return Response({"Message": "OTP is sent plese check your mail or Retry after few seconds"})
    
    if otp_sent:
        otp_sent.delete()
    
    # if otp_sent and otp_sent.is_expire():
    #     otp_sent.delete()
    
    otp = random.randint(100000, 999999)
    
    OTP.objects.create(email = email, otp = otp)
    
    # send_email(email, otp)
    threading.Thread(target=threaded_send_email, args=(email, otp)).start()
    
    return Response({"Message": "Email send Succesfully"})

@api_view(["GET"])
def verify_mail(request):
    email = request.GET.get("email")
    otp = request.GET.get("otp")
    
    try:
        validate_email(email)
        stored_otp = OTP.objects.get(email = email)
        
    except ValidationError:
        return Response({"Message": "Invalid Email"})
        
    except OTP.DoesNotExist:
        return Response({"Message": "Invalid Email"})
    
    if stored_otp.is_expire():
        return Response({"Message": "OTP is Expired"})
    
    if stored_otp.otp != otp:
        return Response({"Message": "Invalid OTP."})
    
    stored_otp.is_verified = True
    stored_otp.save()
    
    return Response({"Message": "OTP verified Succesfully."})


@api_view(['GET'])
@authentication_classes([ExpireTokenAuthentication])
@permission_classes([IsAuthenticated])
def get_users(request):
    user = request.user
    custom_user = CustomUser.objects.get(username = user.username)
    
    user_data = {
        "username": custom_user.username,
        "first name": custom_user.first_name,
        "last name": custom_user.last_name,
        "phone number": custom_user.phone_number,
    }
    
    return Response(user_data)


@api_view(['PATCH'])
@authentication_classes([ExpireTokenAuthentication])
@permission_classes([IsAuthenticated])
def update_user(request):
    data = json.loads(request.body)
    
    email = request.user.email
    username = data['username']
    
    data = json.loads(request.body)
    
    if CustomUser.objects.filter(username = username).exists():
        return Response({'Error': "username is Already Exist"})
    
    custom_user = CustomUser.objects.get(email = email)
    
    if data['username']:
        custom_user.username = data['username']
        
    if data['last_name']:
        custom_user.last_name = data['last_name']
        
    if data['first_name']:
        custom_user.first_name = data['first_name']
    
    custom_user.save()
    
    return Response({'message': 'User updated succesfully'})


@api_view(['DELETE'])
@authentication_classes([ExpireTokenAuthentication])
@permission_classes([IsAuthenticated])
def delete_user(request):
    
    token  = request.auth
    id = request.user.id
    
    # if token.is_expire():
    #     return Response({"message": "Token Expired"})

    custom_user = CustomUser.objects.get(id = id)
    custom_user.delete()
    
    return Response({'message': 'User deleted succesfully'})


# @api_view(['GET'])
# def list_groups_with_users(request):
#     groups = CustomUserRole.objects.all()
#     data = []
    
#     for group in groups:
#         users = group.user_set.all()
#         user_list = []

#         for user in users:
#             user_list.append({
#                 "id": user.id,
#                 "username": user.username,
#                 "email": user.email,
#                 "phone_number": getattr(user, 'phone_number', None)
#             })

#         data.append({
#             "id": group.id,
#             "name": group.name,
#             "description": group.description,
#             "users": user_list
#         })

#     return Response(data)


@api_view(['GET'])
def list_groups_with_users(request):
    user_role = request.GET.get("group_name")
    
    try:
        group = CustomUserRole.objects.get(name = user_role)
    except CustomUserRole.DoesNotExist:
        return Response({"message": "Group name not found"})
    
    # if not group.exists():
    #     return Response({"message": "Group name not found"})

    users = group.user_set.all()
    user_list = []

    for user in users:
        user_list.append({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "phone_number": getattr(user, 'phone_number', None)
        })

    data = {
        "id": group.id,
        "name": group.name,
        "description": group.description,
        "users": user_list
    }

    return Response(data)


##  Group CURD APIs

# class CustomUserRoleListCreateView(APIView):
#     def get(self, request):
#         roles = CustomUserRole.objects.all()
#         serializer = CustomUserRoleSerializer(roles, many=True)
        
#         return Response(serializer.data)

#     def post(self, request):
#         serializer = CustomUserRoleSerializer(data=request.data)
        
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
        
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# class CustomUserRoleDetailView(APIView):
    
#     def get(self, request, pk):
#         role = get_object_or_404(CustomUserRole, pk=pk)
#         serializer = CustomUserRoleSerializer(role)
        
#         return Response(serializer.data)

#     def put(self, request, pk):
#         role = get_object_or_404(CustomUserRole, pk=pk)
#         serializer = CustomUserRoleSerializer(role, data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#     def delete(self, request, pk):
#         role = get_object_or_404(CustomUserRole, pk=pk)
#         role.delete()
#         return Response(status=status.HTTP_204_NO_CONTENT)


# class CustomUserRoleListCreateView(APIView):

#     def get(self, request):
#         roles = CustomUserRole.objects.all()
        
#         data = []
        
#         for role in roles:
#             dct = {
#                 "id": role.id,
#                 "name": role.name,
#                 "description": role.description,
#                 "permissions": [perm.codename for perm in role.permissions.all()]
#             }
            
#             data.append(dct)
            
#         # data = [model_to_dict(role, fields=["id", "name", "description"]) for role in roles]
        
#         return Response(data, status=status.HTTP_200_OK)

#     def post(self, request):
#         name = request.data.get("name")
#         description = request.data.get("description", "")

#         if not name:
#             return Response({"error": "Name is required."}, status = status.HTTP_400_BAD_REQUEST)

#         try:
#             role = CustomUserRole.objects.create(name = name, description = description)
#             return Response(model_to_dict(role), status = status.HTTP_201_CREATED)
        
#         except Exception as e:
#             return Response({"error": str(e)}, status = status.HTTP_400_BAD_REQUEST)


# ## Group Apis 

class CustomUserRoleAPIView(APIView):
    def get(self, request, pk = None):
        if pk:
            role = get_object_or_404(CustomUserRole, pk = pk)
            
            for perm in role.permissions.all():
                permissions.append(perm.codename)
                    
            data = {
                "id": role.id,
                "name": role.name,
                "description": role.description,
                "permissions": permissions
            }
            
            return Response(data=data, status=status.HTTP_200_OK)
        
        else:
            roles = CustomUserRole.objects.all()
                    
            search = request.query_params.get('search')
            
            if search:
                roles = roles.filter(Q(name__icontains = search) | Q(description__icontains = search))

            name_filter = request.query_params.get('name')
            
            if name_filter:
                roles = roles.filter(name = name_filter)

            ordering = request.query_params.get('ordering')
            
            if ordering:
                roles = roles.order_by(ordering)

            data = []
            
            for role in roles:
                permissions = []
                
                for perm in role.permissions.all():
                    permissions.append(perm.codename)
                    
                dct = {
                    "id": role.id,
                    "name": role.name,
                    "description": role.description,
                    "permissions": permissions
                }
                
                data.append(dct)
                
            return Response(data, status = status.HTTP_200_OK)
        
    def post(self, request):
        name = request.data.get("name")
        description = request.data.get("description", "")

        if not name:
            return Response({"error": "Name is required."}, status = status.HTTP_400_BAD_REQUEST)

        try:
            role = CustomUserRole.objects.create(name = name, description = description)
            
            return Response(model_to_dict(role), status = status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response({"error": str(e)}, status = status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        role = get_object_or_404(CustomUserRole, pk = pk)
        
        name = request.data.get("name")
        description = request.data.get("description")

        if name:
            role.name = name
            
        if  description:
            role.description = description

        try:
            role.save()
            
            return Response({"message": "Role updated succesfully"})
        
        except Exception as e:
            return Response({"error": str(e)})

    def delete(self, request, pk):
        role = get_object_or_404(CustomUserRole, pk = pk)
        role.delete()
        
        return Response({"message": "Deleted successfully"})


# Product 
class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter
    ]
    
    filterset_fields = ['is_active']
    search_fields = ['name', 'description', 'tag']
    ordering_fields = ['name', 'price']
    ordering = ['name']  




@api_view(['POST'])
@authentication_classes([ExpireTokenAuthentication])
@permission_classes([IsAuthenticated])
def place_order(request):
    user = request.user

    try:
        customer = CustomUser.objects.get(username = user.username)

    except CustomUser.DoesNotExist:
        return Response({"error": "Customer not found"}, status = status.HTTP_404_NOT_FOUND)

    items = request.data.get("items", [])
    if not items:
        return Response({"error": "No items provided"}, status = status.HTTP_400_BAD_REQUEST)
    
    for item in items:
        quantity = int(item.get("quantity"))
        product_id = item.get("product_id")

        if not product_id or quantity <= 0:
            return Response({"error": "Invalid product ID or quantity"}, status = 400)
        
        try:
            # product = Product.objects.get(id = product_id)
            
            product = Product.objects.filter(id = product_id).first()
            
            if not product:
                return Response({"error": f"Product with ID {product_id} is Not found"}, status = 404)

            if not product.is_active:
                return Response({"error": f"Product with ID {product_id} is not Active."})
                
        except Product.DoesNotExist:
            return Response({"error": f"Product with ID {product_id} is Not found"}, status = 404)

    order = Order.objects.create(customer = customer)
    total = 0.00

    for item in items:
        product_id = item.get("product_id")
        quantity = int(item.get("quantity", 1))

        product = Product.objects.get(id = product_id, is_active = True)
        price = product.price
        subtotal = float(price) * quantity
        total = subtotal + total

        OrderItem.objects.create(
            order = order,
            product = product,
            quantity = quantity,
            price = price
        )

    order.total_amount = total
    order.save()

    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        "orders",
        {
            "type": "send_order_notification",
            "data": {
                "message": f"New order placed by {customer.username} (Order ID: {order.id})"
            }
        }
    )

    Payment.objects.create(
        order = order,
        amount = total,
        status="PENDING",
    )

    return Response({
        "message": "Order placed successfully",
        "order_id": order.id,
        "total_amount": str(total)
        }, status=201)




@api_view(['POST'])
@authentication_classes([ExpireTokenAuthentication])
@permission_classes([IsAuthenticated])
def make_payment(request):

    user = request.user

    order_id = request.data.get("order_id")
    payment_mode = request.data.get("payment_mode", "Cash")
    transaction_id = request.data.get("transaction_id", None)

    valid_modes = dict(Payment.PAYMENT_MODE).keys()
    
    if payment_mode not in valid_modes:
        return Response({"error": "Invalid payment mode"}, status = status.HTTP_400_BAD_REQUEST)

    try:
        order = Order.objects.get(id = order_id)
        payment = Payment.objects.get(order = order)

        if payment.status == "COMPLETED":
            return Response({"message": "Payment already completed for this order."}, status = status.HTTP_200_OK)

        payment.status = "COMPLETED"
        payment.payment_mode = payment_mode
        payment.transaction_id = transaction_id
        payment.save()

        return Response({
            "message": "Payment completed successfully",
            "order_id": order.id,
            "amount": str(payment.amount),
            "payment_mode": payment.payment_mode,
            "transaction_id": payment.transaction_id,
            "status": payment.status
        }, status = 200)

    except Order.DoesNotExist:
        return Response({"error": "Order not found"}, status = status.HTTP_404_NOT_FOUND)

    except Payment.DoesNotExist:
        return Response({"error": "Payment record not found for this order"}, status = status.HTTP_404_NOT_FOUND)


# @api_view(['GET'])
# @authentication_classes([ExpireTokenAuthentication])
# @permission_classes([IsAuthenticated])
# def view_orders_paginated(request):
#     user = request.user

#     try:
#         customer = CustomUser.objects.get(username = user.username)
        
#     except CustomUser.DoesNotExist:
#         return Response({"error": "Customer not found"}, status=status.HTTP_404_NOT_FOUND)

#     page_size = request.GET.get('size', 5)
    
#     try:
#         page_size = int(page_size)
        
#         if page_size <= 0:
#             raise ValueError
        
#     except ValueError:
#         return Response({"error": "Invalid page size"}, status = status.HTTP_400_BAD_REQUEST)

#     paginator = PageNumberPagination()
#     paginator.page_size = page_size

#     orders = Order.objects.filter(customer = customer)
#     result_page = paginator.paginate_queryset(orders, request)

#     orders_data = []
    
#     for order in result_page:
#         items = OrderItem.objects.filter(order = order)
#         item_list = []
        
#         for item in items:
#             item_list.append({
#                 "product": item.product.name,
#                 "quantity": item.quantity,
#                 "price": str(item.price),
#                 "subtotal": str(item.price * item.quantity),
#             })

#         orders_data.append({
#             "order_id": order.id,
#             "total_amount": str(order.total_amount),
#             "items": item_list
#         })

#     return paginator.get_paginated_response(orders_data)



from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

@api_view(['GET'])
@authentication_classes([ExpireTokenAuthentication])
@permission_classes([IsAuthenticated])
def view_orders_paginated(request):
    user = request.user

    try:
        customer = CustomUser.objects.get(username=user.username)
    except CustomUser.DoesNotExist:
        return Response({"error": "Customer not found"}, status=status.HTTP_404_NOT_FOUND)

    # Get page number and size from query params
    page_number = request.GET.get('page', 1)
    page_size = request.GET.get('size', 5)

    try:
        page_number = int(page_number)
        page_size = int(page_size)
        if page_number <= 0 or page_size <= 0:
            raise ValueError
    except ValueError:
        return Response({"error": "Invalid page number or page size"}, status=status.HTTP_400_BAD_REQUEST)

    orders = Order.objects.filter(customer=customer).order_by('-id')  # optional ordering

    paginator = Paginator(orders, page_size)

    try:
        page = paginator.page(page_number)
    except PageNotAnInteger:
        return Response({"error": "Page number is not an integer"}, status=status.HTTP_400_BAD_REQUEST)
    except EmptyPage:
        return Response({"error": "Page out of range"}, status=status.HTTP_404_NOT_FOUND)

    orders_data = []
    for order in page:
        items = OrderItem.objects.filter(order=order)
        item_list = []
        for item in items:
            item_list.append({
                "product": item.product.name,
                "quantity": item.quantity,
                "price": str(item.price),
                "subtotal": str(item.price * item.quantity),
            })

        orders_data.append({
            "order_id": order.id,
            "total_amount": str(order.total_amount),
            "items": item_list
        })

    response_data = {
        "current_page": page_number,
        "total_pages": paginator.num_pages,
        "next_page": page_number + 1 if page.has_next() else None,
        "previous_page": page_number - 1 if page.has_previous() else None,
        "total_orders": paginator.count,
        "orders": orders_data
    }

    return Response(response_data, status=status.HTTP_200_OK)




# #  Download producs in excel file
# @api_view(['GET'])
# def download_products_excel(request):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Products"
#     ws.append(['ID', 'Name', 'Description', 'Price', 'Tag', 'Category', 'Is Active'])

#     products = Product.objects.select_related('category').all()
#     for product in products:
#         ws.append([
#             product.id,
#             smart_str(product.name),
#             smart_str(product.description),
#             float(product.price),
#             product.tag,
#             smart_str(product.category.name),
#             product.is_active,
#         ])

#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=products.xlsx'

#     wb.save(response)
#     return response


@api_view(['GET'])
def download_products_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(['Name', 'Description', 'Price', 'Tag', 'Category', 'Is Active'])

    products = Product.objects.select_related('category').all()
    for product in products:
        ws.append([
            smart_str(product.name),
            smart_str(product.description),
            float(product.price),
            product.tag,
            smart_str(product.category.name),
            product.is_active,
        ])

    file_name = "products.xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)
    wb.save(file_path)

    file_url = request.build_absolute_uri(settings.MEDIA_URL + file_name)

    return Response({
        "message": "Excel file generated successfully.",
        "file_url": file_url
    })


# @api_view(['POST'])
# @parser_classes([MultiPartParser])
# def upload_products_excel(request):
#     file = request.FILES.get('file')

#     if not file:
#         return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

#     try:
#         wb = load_workbook(file)
#         ws = wb.active
        
#     except Exception as e:
#         return Response({"error": f"Invalid Excel file. {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

#     rows = list(ws.iter_rows(values_only=True))
#     if not rows or rows[0] != ('Name', 'Description', 'Price', 'Tag', 'Category', 'Is Active'):
#         return Response({"error": "Invalid header format."}, status=status.HTTP_400_BAD_REQUEST)

#     success_count = 0
#     not_uploadeds = []
#     failed_rows = []
    
#     for row in rows[1:]:
#         try:
#             name, description, price, tag, category_name, is_active = row
#             print(row)
            
#             if not name:
#                 continue

#             category = Category.objects.filter(name=category_name).first()

#             if not category:
#                 not_uploadeds.append(name)

#             Product.objects.update_or_create(
#                 name = name,
#                 defaults={
#                     'description': description,
#                     'price': price,
#                     'tag': tag,
#                     'category': category,
#                     'is_active': parse_boolean(is_active),
#                 }
#             )
#             success_count = success_count + 1

#         except Exception as e:
#             failed_rows.append([
#                 row[0], row[1], row[2], row[3], row[4], row[5], str(e)
#             ])
        
#     response_data = {
#     "message": f"{success_count} products uploaded successfully.",
#     } 
        
#     if not_uploadeds:
#         response_data["not_uploaded_products"] = {
#             "reason": "Missing category in database.",
#             "products": not_uploadeds
#         }
    

#     return Response(response_data)





# from rest_framework.decorators import api_view
# from rest_framework.response import Response
# from rest_framework import status
# from openpyxl import load_workbook, Workbook
# from django.http import HttpResponse
# from django.utils.encoding import smart_str
# from io import BytesIO
# from .models import Product, Category

# def parse_boolean(value):
#     return str(value).strip().lower() in ['1', 'true', 'yes']

# def sanitize_string(value):
#     return str(value).strip() if value is not None else ''

# def sanitize_price(value):
#     try:
#         return float(value)
#     except:
#         return None

# @api_view(['POST'])
# def upload_products_excel(request):
#     file = request.FILES.get('file')

#     if not file:
#         return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

#     try:
#         wb = load_workbook(file)
#         ws = wb.active
#     except Exception as e:
#         return Response({"error": f"Invalid Excel file. {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

#     rows = list(ws.iter_rows(values_only=True))
#     if not rows or rows[0] != ('Name', 'Description', 'Price', 'Tag', 'Category', 'Is Active'):
#         return Response({"error": "Invalid header format."}, status=status.HTTP_400_BAD_REQUEST)

#     success_count = 0
#     failed_rows = []

#     for idx, row in enumerate(rows[1:], start=2):
#         try:
#             name = sanitize_string(row[0])
#             description = sanitize_string(row[1])
#             price = sanitize_price(row[2])
#             tag = sanitize_string(row[3])
#             category_name = sanitize_string(row[4])
#             is_active = parse_boolean(row[5])

#             if not name:
#                 raise ValueError("Product name is missing.")

#             if price is None:
#                 raise ValueError("Invalid price format.")

#             category = Category.objects.filter(name=category_name).first()
            
#             if not category:
#                 raise ValueError(f"Category '{category_name}' not found.")

#             Product.objects.update_or_create(
#                 name=name,
#                 defaults={
#                     'description': description,
#                     'price': price,
#                     'tag': tag,
#                     'category': category,
#                     'is_active': is_active,
#                 }
#             )
#             success_count += 1

#         except Exception as e:
#             failed_rows.append([
#                 row[0], row[1], row[2], row[3], row[4], row[5], str(e)
#             ])

#     if failed_rows:
#         error_wb = Workbook()
#         error_ws = error_wb.active
#         error_ws.title = "Failed Uploads"
#         error_ws.append(['Name', 'Description', 'Price', 'Tag', 'Category', 'Is Active', 'Error'])

#         for failed_row in failed_rows:
#             error_ws.append(failed_row)

#         output = BytesIO()
#         error_wb.save(output)
#         output.seek(0)
        
#         file_name = "failed_products.xlsx"
#         file_path = os.path.join(settings.MEDIA_ROOT, file_name)
#         wb.save(file_path)

#         response = HttpResponse(
#             output,
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = 'attachment; filename=failed_products.xlsx'
#         return response
    

#     return Response({
#         "message": f"{success_count} products uploaded successfully.",
#     })



def sanitize_string(value):
    if value:
        return str(value).strip()
    else :
        return  ''

def sanitize_price(value):
    try:
        return float(value)
    except:
        return None
    
def sanitize_tag(tag):
    if tag in ("Veg", "VEG", "veg"):
        return str('VEG')
    
    elif tag in ("Non-Veg", "Non Veg", "non-veg", 'NON-VEG'):
        return str('NON-VEG')
    
    else :
        return None
    
@api_view(['POST'])
def upload_products_excel(request):
    file = request.FILES.get('file')

    if not file:
        return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = load_workbook(file)
        ws = wb.active
    except Exception as e:
        return Response({"error": f"Invalid Excel file. {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    rows = list(ws.iter_rows(values_only=True))
    if not rows or rows[0] != ('Name', 'Description', 'Price', 'Tag', 'Category', 'Is Active'):
        return Response({"error": "Invalid header format."}, status=status.HTTP_400_BAD_REQUEST)

    success_count = 0
    failed_rows = []

    for row in rows[1:]:
        try:

            name = row[0]
            description = sanitize_string(row[1])
            price = row[2]
            tag = row[3]
            category_name = row[4]
            is_active = row[5]

            # if not row[2]:
            #     raise ValueError("Price is missing")
            
            # if not name:
            #     raise ValueError("Product name is missing.")
            
            # if len(name) > 50:
            #     raise ValueError("More than 50 charecter of name not alllowed")
            
            # prod = Product.objects.filter(name = name).first()
            
            # if prod:
            #     raise ValueError("Product is already present with this Name.")
        
            # if price is None:
            #     raise ValueError("Invalid price format.")
            
            # if price > 99999999.99:
            #     raise ValueError("Price Must be striclty less than 100000000.00")

            # category = Category.objects.filter(name=category_name).first()
            
            # if not category:
            #     raise ValueError(f"Category '{category_name}' not found.")
            
            # if not tag: 
            #     raise ValueError("Invalid Tag")
            
            
            error = []

            ## Name
            if not name:
                error.append("Product name is missing.")

            if name and len(name) > 50:
                error.append("More than 50 characters in name not allowed.")

            if name:
                prod = Product.objects.filter(name=name).first()
                if prod:
                    error.append("Product is already present with this name.")
                
            ## Price        
            if not row[2]:
                error.append("Price is missing")
                
            elif sanitize_price(price) is None:
                error.append(" Invalid Price Format")
                
            elif float(price) > 99999999.99:
                error.append("Price must be strictly less than 100000000.00")
                
            else:
                price = float(price)
                
            ## Category
            if not category_name:
                error.append("Category is Missing")
            else:
                category = Category.objects.filter(name = category_name).first()
                
                if not category:
                    error.append(f"Category '{category_name}' not found.")

            ## Tag 
            if not tag:
                error.append('Tag is Missing and Tag must be "VEG" or "NON-VEG"')
                
            elif tag not in ("VEG", "NON-VEG"):
                error.append('Tag must be "VEG" or "NON-VEG"')
                
            if not is_active:
                error.append("Is_active field is missing")
            
            else:
                is_active = parse_boolean(is_active)
        
                if not isinstance(is_active, bool):
                    error.append("Invalid is_active field it must be TRUE of FALSE")
                    
            if not error:
                Product.objects.create(
                    name = name,
                    description = description,
                    price = price,
                    tag = tag,
                    category = category,
                    is_active = is_active,
                )
                
                success_count = success_count + 1
                
            else:
                failed_rows.append([
                    row[0], row[1], row[2], row[3], row[4], row[5], str(", ".join(error))
                ])
                
        except Exception as e:
            return Response({"error": f"Internl server error as {e}"}, status = 500)

    response_data = {
        "message": f"{success_count} products uploaded successfully.",
    }

    if failed_rows:
        error_wb = Workbook()
        error_ws = error_wb.active
        error_ws.title = "Failed Uploads"
        error_ws.append(['Name', 'Description', 'Price', 'Tag', 'Category', 'Is Active', 'Errors'])

        for failed_row in failed_rows:
            error_ws.append(failed_row)

        output = BytesIO()
        error_wb.save(output)
        output.seek(0)

        filename = f"failed_products.xlsx"
        media_path = os.path.join("failed_uploads", filename)
        full_path = default_storage.save(media_path, ContentFile(output.read()))
        file_url = request.build_absolute_uri(os.path.join(settings.MEDIA_URL, full_path))

        response_data["failed_uploads_file_url"] = file_url
        response_data["failed_count"] = len(failed_rows)

    return Response(response_data)


# In-memory store (replace with DB model in production)
device_tokens = set()

@api_view(['POST'])
def register_token(request):
    token = request.data.get("token")
    print(token)
    if token:
        device_tokens.add(token)
        return Response({"message": "Token registered"}, status=status.HTTP_201_CREATED)
    return Response({"error": "Token is required"}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def send_notification(request):
    from .utils import send_push_notification
    success = 0

    for token in device_tokens:
        
        try:
            send_push_notification(token, "Hello!", "You got a test notification 🎉")
            success += 1
            
        except Exception as e:
            print(f"Error sending to {token}: {e}")
            
    return Response({"message": f"Notification sent to {success} devices."}, status=status.HTTP_200_OK)



@api_view(['DELETE'])
@authentication_classes([ExpireTokenAuthentication])
@permission_classes([IsAdminUser])
def delete_expired_users(request):
    current_time = timezone.now()
    
    expired_users = CustomUser.objects.filter(expire_Date__lt=current_time)
    usernames = list(expired_users.values_list('username', flat=True))

    expired_users.delete()

    return Response(
        {
            "message": f"{len(usernames)} expired user(s) deleted.",
            "deleted_usernames": usernames
        },
        status=status.HTTP_200_OK
    )
    
    